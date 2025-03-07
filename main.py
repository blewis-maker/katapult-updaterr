import json
import http.client
import geopandas as gpd
from shapely.geometry import Point, LineString
from shapely.geometry import mapping
import os
import time
import socket
import re
import pandas as pd
from datetime import datetime
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl import load_workbook
from openpyxl.cell import MergedCell
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
import logging
from dotenv import load_dotenv
import msal
from O365 import Account, FileSystemTokenBackend
import requests
import base64
import warnings
import zipfile
from io import BytesIO
import traceback
from flask import Flask, request, jsonify
warnings.filterwarnings('ignore')

# Add to imports at the top
from arcgis_updater import ArcGISUpdater

app = Flask(__name__)

# Add at start of script
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[logging.StreamHandler()]
)

# Load environment variables from .env file
load_dotenv()

# Toggle to enable/disable testing a specific job
TEST_ONLY_SPECIFIC_JOB = False

# IDs of the test jobs
TEST_JOB_IDS = [
    "-O-nlOLQbPIYhHwJCPDN",
    "-Nvs8uA2MHZB5NdTK2_p",
    "-O4RHaixdJmN3lqi0Q3m",
    "-O7_Gr-6exIhw0vtfVga"
]

# Add at top of script
CONFIG = {
    'API_KEY': 'rt2JR8Rds03Ry03hQTpD9j0N01gWEULJnuY3l1_GeXA8uqUVLtXsKHUQuW5ra0lt-FklrA40qq6_J04yY0nPjlfKG1uPerclUX2gf6axkIioJadYxzOG3cPZJLRcZ2_vHPdipZWvQdICAL2zRnqnOUCGjfq4Q8aMdmA7H6z7xK7W9MEKnIiEALokmtChLtr-s6hDFko17M7xihPpNlfGN7N8D___wn55epkLMtS2eFF3JPlj_SjpFIGXYK15PJFta-BmPqCFvEwXlZEYfEf8uYOpAvCEdBn3NOMoB-P28owOJ7ZeBQf5VMFi3J5_RV2fE_XDR2LTD469Qq0y3946LQ',
    'WORKSPACE_PATH': os.path.expanduser('~/reverseengineerAPI/reverse_engineer_API/workspace'),
    'SHAREPOINT': {
        'SITE_URL': 'deeplydigital.sharepoint.com:/sites/OSPIntegrationTestingSite',  # Updated to new site
        'DRIVE_PATH': 'Documents',  # Simplified path for testing
        'FILE_NAME': 'Aerial_Status_Tracker.xlsx'  # Updated file name
    }
}

# Email configuration
EMAIL_CONFIG = {
    'client_id': os.getenv('AZURE_CLIENT_ID'),
    'client_secret': os.getenv('AZURE_CLIENT_SECRET'),
    'tenant_id': os.getenv('AZURE_TENANT_ID'),
    'user_email': 'brandan.lewis@deeplydigital.com',
    'default_recipients': ['brandan.lewis@deeplydigital.com']
}

# Function to get list of jobs from KatapultPro API
def getJobList():
    URL_PATH = '/api/v2/jobs'
    headers = {}
    all_jobs = []
    base_wait_time = 2  # Reduced from 5 to 2 seconds
    
    for attempt in range(5):
        conn = None
        try:
            # More efficient backoff calculation
            wait_time = min(base_wait_time * (1.5 ** attempt), 15)  # Cap at 15 seconds
            if attempt > 0:
                time.sleep(wait_time)
            
            conn = http.client.HTTPSConnection("katapultpro.com", timeout=30)
            
            # Print the full URL being requested
            full_url = f"{URL_PATH}?api_key={CONFIG['API_KEY'][:10]}..."  # Only show first 10 chars of API key
            print(f"Requesting URL: https://katapultpro.com{full_url}")
            
            conn.request("GET", f"{URL_PATH}?api_key={CONFIG['API_KEY']}", headers=headers)
            res = conn.getresponse()
            
            # Print response status and headers
            print(f"Response Status: {res.status} {res.reason}")
            
            data = res.read().decode("utf-8")
            
            if res.status == 429:  # Too Many Requests
                print("Rate limit exceeded. Will retry with exponential backoff.")
                continue
            
            if res.status != 200:
                print(f"Error: Received status code {res.status}")
                print(f"Response: {data[:500]}")
                continue

            try:
                jobs_dict = json.loads(data)
                
                # Check if jobs_dict is a string and try to parse it again if needed
                if isinstance(jobs_dict, str):
                    jobs_dict = json.loads(jobs_dict)

                if not isinstance(jobs_dict, dict):
                    print(f"Unexpected response format: {type(jobs_dict)}")
                    print(f"Response content: {jobs_dict}")
                    continue

                # Process each job to get the correct name
                all_jobs = []
                for job_id, job_details in jobs_dict.items():
                    if isinstance(job_details, dict):
                        # Try to get name from metadata first
                        metadata = job_details.get('metadata', {})
                        job_name = metadata.get('name')
                        
                        # If no name in metadata, try job_details directly
                        if not job_name:
                            job_name = job_details.get('name')
                        
                        # If still no name, use job ID
                        if not job_name:
                            job_name = f"Job {job_id}"
                            
                        job_status = metadata.get('job_status', 'Unknown')
                        all_jobs.append({
                            'id': job_id,
                            'name': job_name,
                            'status': job_status
                        })
                
                if all_jobs:
                    print(f"Successfully retrieved {len(all_jobs)} jobs")
                    break
                else:
                    print("No jobs found in the response")
                    print(f"Raw response: {data[:500]}")
                    continue

            except json.JSONDecodeError as e:
                print(f"JSON decode error: {e}")
                print(f"Raw data causing the error: {data[:200]}...")
                continue

        except (socket.error, OSError) as e:
            print(f"Socket error: {e}. Will retry...")
        except Exception as e:
            print(f"Failed to retrieve job list: {str(e)}")
            if 'data' in locals():
                print(f"Response data: {data[:200]}...")
        finally:
            if conn:
                try:
                    conn.close()
                except:
                    pass

    return all_jobs

# Function to get job data from KatapultPro API
def getJobData(job_id):
    URL_PATH = f'/api/v2/jobs/{job_id}'
    headers = {}
    job_data = None  # Initialize job_data to None
    max_retries = 5
    base_timeout = 30  # Reduced from 60 to 30 seconds
    
    for attempt in range(max_retries):
        conn = None
        current_timeout = min(base_timeout * (1.5 ** attempt), 45)  # Cap timeout at 45 seconds
        wait_time = min(2 * (1.5 ** attempt), 15)  # More efficient wait time progression
        
        try:
            print(f"Attempt {attempt + 1}/{max_retries} for job {job_id} (timeout: {current_timeout}s)")
            conn = http.client.HTTPSConnection("katapultpro.com", timeout=current_timeout)
            conn.request("GET", f"{URL_PATH}?api_key={CONFIG['API_KEY']}", headers=headers)
            res = conn.getresponse()
            
            # Check response status
            if res.status != 200:
                print(f"Received status code {res.status} for job {job_id}")
                if res.status == 429:  # Rate limit exceeded
                    print(f"Rate limit exceeded. Waiting {wait_time} seconds before retry...")
                    time.sleep(wait_time)
                    continue
                elif res.status >= 500:  # Server error
                    print(f"Server error. Waiting {wait_time} seconds before retry...")
                    time.sleep(wait_time)
                    continue
            
            data = res.read().decode("utf-8")
            job_data = json.loads(data)

            if "error" in job_data:
                if job_data["error"] == "RATE LIMIT EXCEEDED":
                    print(f"Rate limit exceeded. Waiting {wait_time} seconds before retry...")
                    time.sleep(wait_time)
                    continue
                else:
                    print(f"API error: {job_data['error']}")
                    time.sleep(wait_time)
                    continue

            # Save job data to a file if testing a specific job
            if TEST_ONLY_SPECIFIC_JOB:
                workspace_path = CONFIG['WORKSPACE_PATH']
                file_path = os.path.join(workspace_path, f"test_job_{job_id.replace('/', '_')}.json")
                with open(file_path, 'w') as f:
                    json.dump(job_data, f, indent=2)
                print(f"Job data saved to: {file_path}")
            
            print(f"Successfully retrieved data for job {job_id}")
            return job_data

        except json.JSONDecodeError:
            print(f"Failed to decode JSON for job {job_id}")
            time.sleep(wait_time)
        except (socket.error, OSError) as e:
            print(f"Socket error while retrieving job data for {job_id}: {e}. Waiting {wait_time} seconds before retry...")
            time.sleep(wait_time)
        except Exception as e:
            print(f"Error retrieving job data for {job_id}: {e}")
            time.sleep(wait_time)
        finally:
            if conn:
                try:
                    conn.close()
                except:
                    pass
            time.sleep(1)  # Small delay between attempts

    print(f"Failed to retrieve job data for job ID {job_id} after {max_retries} attempts.")
    if job_data:  # Only print job_data if it exists
        print(f"Last received Job Data for {job_id}:\n{json.dumps(job_data, indent=2)}")
    return None  # Return None if all attempts fail


# Extract nodes (poles, anchors, etc.) from job data
def extractNodes(job_data, job_name, job_id, user_map):
    nodes = job_data.get("nodes", {})
    if not nodes:
        print("No nodes found.")
        return []

    # Analysis counters (keep existing counters)
    node_type_counts = {}
    attachment_height_counts = {}
    pole_class_counts = {}
    pole_height_counts = {}
    
    # Debug counters
    total_nodes = len(nodes)
    processed_nodes = 0
    skipped_nodes = 0
    nodes_with_height = 0
    nodes_without_height = 0
    
    print(f"\nStarting node analysis for {total_nodes} total nodes...")
    print("------------------------")
    
    photo_data = job_data.get('photos', {})
    trace_data_all = job_data.get('traces', {}).get('trace_data', {})
    node_points = []

    # Extract job status and conversation from job data
    metadata = job_data.get('metadata', {})
    job_status = metadata.get('job_status', "Unknown")
    conversation = metadata.get('conversation', "")
    project = metadata.get('project', "")

    for node_id, node_data in nodes.items():
        try:
            attributes = node_data.get('attributes', {})
            
            # Extract editor tracking information
            editors_history = {}  # Changed to dict to track latest edit per editor
            last_editor = "Unknown"
            last_edit_time = None
            
            # Check photos associated with the node for editor information
            node_photos = node_data.get('photos', {})
            for photo_id in node_photos:
                if photo_id in photo_data:
                    photo_editors = photo_data[photo_id].get('photofirst_data', {}).get('_editors', {})
                    if photo_editors:
                        # Process all editors for this photo
                        for editor_id, edit_time in photo_editors.items():
                            editor_name = user_map.get(editor_id, "Unknown User")
                            # Only update if this is the most recent edit for this editor
                            if editor_name not in editors_history or edit_time > editors_history[editor_name]['raw_timestamp']:
                                # Convert timestamp to MST (UTC-7)
                                edit_dt = datetime.fromtimestamp(edit_time/1000)
                                # Subtract 7 hours for MST
                                edit_dt_mst = edit_dt.replace(hour=(edit_dt.hour - 7) % 24)
                                formatted_time = edit_dt_mst.strftime('%I:%M %p MST')
                                formatted_date = edit_dt_mst.strftime('%Y-%m-%d')
                                
                                editors_history[editor_name] = {
                                    'editor': editor_name,
                                    'timestamp': f"{formatted_date} {formatted_time}",
                                    'raw_timestamp': edit_time
                                }

            # Convert editors_history dict to list and sort by timestamp
            editors_list = list(editors_history.values())
            if editors_list:
                editors_list.sort(key=lambda x: x['raw_timestamp'], reverse=True)
                last_editor = editors_list[0]['editor']
                last_edit_time = editors_list[0]['timestamp']
                
                # Print detailed editor history for debugging
                print(f"\nEditor history for node {node_id}:")
                for edit in editors_list:
                    print(f"  Editor: {edit['editor']}")
                    print(f"  Last Edit: {edit['timestamp']}")
                print("------------------------")
            
            # Check if node is a pole
            node_type = 'Unknown'
            for type_source in ['node_type', 'pole_type']:
                for source_type in ['-Imported', 'button_added', 'value', 'auto_calced']:
                    type_value = attributes.get(type_source, {}).get(source_type)
                    if type_value:
                        node_type = type_value
                        break
                if node_type != 'Unknown':
                    break
            
            # Count node types
            if node_type not in node_type_counts:
                node_type_counts[node_type] = 0
            node_type_counts[node_type] += 1
            
            # Only process poles (exclude anchors, references, etc.)
            if node_type == 'pole':
                processed_nodes += 1
                latitude = node_data.get('latitude')
                longitude = node_data.get('longitude')

                if latitude is None or longitude is None:
                    print(f"Warning: Missing coordinates for pole {node_id}")
                    skipped_nodes += 1
                    continue

                # Extract MR status
                mr_status = "Unknown"
                if 'proposed_pole_spec' in attributes:
                    mr_status = "PCO Required"
                else:
                    mr_state = attributes.get('mr_state', {}).get('auto_calced', "Unknown")
                    warning_present = 'warning' in attributes
                    if mr_state == "No MR" and not warning_present:
                        mr_status = "No MR"
                    elif mr_state == "MR Resolved" and not warning_present:
                        mr_status = "Comm MR"
                    elif mr_state == "MR Resolved" and warning_present:
                        mr_status = "Electric MR"

                # Extract pole attributes
                pole_tag = attributes.get('pole_tag', {})
                
                # First try to get company from -Imported or button_added
                company = pole_tag.get('-Imported', {}).get('company') or pole_tag.get('button_added', {}).get('company')
                
                # If not found, look for company in any direct child of pole_tag
                if not company:
                    for key, value in pole_tag.items():
                        if isinstance(value, dict) and 'company' in value:
                            company = value['company']
                            break
                
                # If still not found, use default
                if not company:
                    company = "Unknown"
                
                fldcompl_value = attributes.get('field_completed', {}).get('value', "Unknown")
                fldcompl = 'yes' if fldcompl_value == 1 else 'no' if fldcompl_value == 2 else 'Unknown'
                
                # Extract and count pole class
                pole_class = attributes.get('pole_class', {}).get('-Imported', "Unknown")
                if pole_class not in pole_class_counts:
                    pole_class_counts[pole_class] = 0
                pole_class_counts[pole_class] += 1
                
                # Extract and count pole height
                pole_height = attributes.get('pole_height', {}).get('-Imported', "Unknown")
                if pole_height not in pole_height_counts:
                    pole_height_counts[pole_height] = 0
                pole_height_counts[pole_height] += 1
                
                # Extract tag and scid
                tag = attributes.get('pole_tag', {}).get('-Imported', {}).get('tagtext', "Unknown")
                scid = attributes.get('scid', {}).get('auto_button', "Unknown")

                # Extract POA height with detailed analysis
                poa_height = ""
                photos = node_data.get('photos', {})
                main_photo_id = next(
                    (photo_id for photo_id, photo_info in photos.items() if photo_info.get('association') == 'main'), None)

                if main_photo_id and main_photo_id in photo_data:
                    # Check wire data
                    photofirst_data = photo_data[main_photo_id].get('photofirst_data', {}).get('wire', {})
                    for wire_info in photofirst_data.values():
                        trace_id = wire_info.get('_trace')
                        trace_data = trace_data_all.get(trace_id, {})

                        if (trace_data.get('company') == 'Clearnetworx' and
                                trace_data.get('proposed', False) and
                                trace_data.get('_trace_type') == 'cable' and
                                trace_data.get('cable_type') == 'Fiber Optic Com'):

                            poa_height = wire_info.get('_measured_height')
                            if poa_height is not None:
                                feet = int(poa_height // 12)
                                inches = int(poa_height % 12)
                                poa_height = f"{feet}' {inches}\""
                                nodes_with_height += 1
                                
                                # Count attachment heights
                                height_key = f"{feet}'{inches}\""
                                if height_key not in attachment_height_counts:
                                    attachment_height_counts[height_key] = 0
                                attachment_height_counts[height_key] += 1
                                break

                    # Check guying data if no POA height found
                    if not poa_height:
                        guying_data = photo_data[main_photo_id].get('photofirst_data', {}).get('guying', {})
                        for wire_info in guying_data.values():
                            trace_id = wire_info.get('_trace')
                            trace_data = trace_data_all.get(trace_id, {})

                            if (trace_data.get('company') == 'Clearnetworx' and
                                    trace_data.get('proposed', False) and
                                    trace_data.get('_trace_type') == 'down_guy'):

                                poa_height = wire_info.get('_measured_height')
                                if poa_height is not None:
                                    feet = int(poa_height // 12)
                                    inches = int(poa_height % 12)
                                    poa_height = f"{feet}' {inches}\""
                                    nodes_with_height += 1
                                    
                                    # Count attachment heights
                                    height_key = f"{feet}'{inches}\""
                                    if height_key not in attachment_height_counts:
                                        attachment_height_counts[height_key] = 0
                                    attachment_height_counts[height_key] += 1
                                    break

                if not poa_height:
                    nodes_without_height += 1

                node_points.append({
                    "node_id": node_id,  # Use the actual node ID from the JSON
                    "lat": latitude,
                    "lng": longitude,
                    "jobname": job_name,
                    "job_status": job_status,
                    "MR_statu": mr_status,
                    "company": company,
                    "fldcompl": fldcompl,
                    "pole_class": pole_class,
                    "tag": tag,
                    "scid": scid,
                    "POA_Height": poa_height,
                    "conversation": conversation,
                    "project": project,
                    "last_editor": last_editor,
                    "last_edit": last_edit_time
                })
        except Exception as e:
            print(f"Error processing node {node_id}: {str(e)}")
            skipped_nodes += 1

    # Print analysis results
    print("\nNode Processing Summary:")
    print(f"Total nodes found: {total_nodes}")
    print(f"Poles processed: {processed_nodes}")
    print(f"Skipped nodes: {skipped_nodes}")
    print(f"Poles with height: {nodes_with_height}")
    print(f"Poles without height: {nodes_without_height}")
    
    print("\nNode Type Distribution:")
    for ntype, count in sorted(node_type_counts.items()):
        print(f"{ntype}: {count}")
    
    print("\nPole Class Distribution:")
    for pclass, count in sorted(pole_class_counts.items()):
        print(f"{pclass}: {count}")
    
    print("\nPole Height Distribution:")
    for pheight, count in sorted(pole_height_counts.items()):
        print(f"{pheight}: {count}")
    
    print("\nAttachment Height Distribution:")
    for height, count in sorted(attachment_height_counts.items()):
        print(f"{height}: {count}")
    
    print("------------------------\n")

    return node_points

def extractAnchors(job_data, job_name, job_id):
    """Extract anchor points from job data, with detailed type analysis."""
    anchors = job_data.get("nodes", {})
    anchor_points = []
    
    # Analysis counters
    node_type_counts = {}
    anchor_spec_counts = {}
    anchor_status = {}
    
    print("\nDetailed Anchor Analysis:")
    print("------------------------")
    
    for node_id, node_data in anchors.items():
        attributes = node_data.get("attributes", {})
        
        # Check all possible node type fields
        node_type = "Unknown"
        for type_field in ["node_type", "anchor_type"]:
            for source in ["button_added", "-Imported", "value", "auto_calced"]:
                type_value = attributes.get(type_field, {}).get(source)
                if type_value and "anchor" in str(type_value).lower():
                    node_type = type_value
                    break
            if node_type != "Unknown":
                break
                
        # Count node types
        if node_type not in node_type_counts:
            node_type_counts[node_type] = 0
        node_type_counts[node_type] += 1
        
        # Only process if it's an anchor
        if "anchor" in str(node_type).lower():
            latitude = node_data.get("latitude")
            longitude = node_data.get("longitude")
            
            # Get anchor specification with detailed logging
            anchor_spec = "Unknown"
            print(f"\nAnalyzing anchor {node_id} (Type: {node_type}):")
            
            # Check anchor_spec field for both multi_added and button_added
            anchor_spec_data = attributes.get("anchor_spec", {})
            if anchor_spec_data.get("multi_added"):
                anchor_spec = anchor_spec_data.get("multi_added")
                print(f"  Found spec in multi_added: {anchor_spec}")
            elif anchor_spec_data.get("button_added"):
                anchor_spec = anchor_spec_data.get("button_added")
                print(f"  Found spec in button_added: {anchor_spec}")
            
            # Count anchor specs
            if anchor_spec not in anchor_spec_counts:
                anchor_spec_counts[anchor_spec] = 0
            anchor_spec_counts[anchor_spec] += 1
            
            # Track anchor status (new vs existing)
            status = "new" if "new" in str(node_type).lower() else "existing" if "existing" in str(node_type).lower() else "unknown"
            if status not in anchor_status:
                anchor_status[status] = 0
            anchor_status[status] += 1
            
            # Append anchor information
            anchor_points.append({
                "longitude": longitude,
                "latitude": latitude,
                "anchor_spec": anchor_spec,
                "anchor_type": node_type,
                "job_id": job_id
            })
    
    # Print summary analysis
    print("\nSummary Analysis:")
    print("------------------------")
    print("Node Types Found:")
    for ntype, count in sorted(node_type_counts.items()):
        print(f"{ntype}: {count}")
    
    print("\nAnchor Specifications:")
    for spec, count in sorted(anchor_spec_counts.items()):
        print(f"{spec}: {count}")
    
    print("\nAnchor Status:")
    for status, count in sorted(anchor_status.items()):
        print(f"{status}: {count}")
    print("------------------------")

    return anchor_points
# Extract connections (lines, cables, etc.) from job data
def extractConnections(connections, nodes, job_data=None):
    """Extract connection data with enhanced wire specification and height information."""
    # Analysis counters
    connection_type_counts = {}
    connection_height_counts = {}
    
    # Debug counters for wire_spec and mid_ht
    total_aerial_cables = 0
    aerial_with_wire_spec = 0
    aerial_with_mid_ht = 0
    
    # Debug counters
    total_connections = len(connections)
    connections_with_height = 0
    connections_without_height = 0
    connections_by_type = {}
    
    print(f"\nStarting connection analysis for {total_connections} total connections...")
    print("------------------------")
    
    # Get trace data and photo data from job data
    trace_data = job_data.get('traces', {}).get('trace_data', {}) if job_data else {}
    photo_data = job_data.get('photos', {}) if job_data else {}
    
    valid_connections = []
    processed_count = 0
    skipped_count = 0
    
    # Store some sample data for logging
    sample_aerial_cables = []
    
    for connection_id, connection_data in connections.items():
        try:
            # Get node IDs from the connection data
            node_id_1 = connection_data.get('node_id_1')  # First try direct node_id_1
            node_id_2 = connection_data.get('node_id_2')  # First try direct node_id_2
            
            # If not found directly, try getting from node_1 and node_2 objects
            if not node_id_1:
                node_id_1 = connection_data.get('node_1', {}).get('id')
            if not node_id_2:
                node_id_2 = connection_data.get('node_2', {}).get('id')
            
            if not node_id_1 or not node_id_2:
                print(f"Warning: Missing node IDs for connection {connection_id} (nodes: {node_id_1}, {node_id_2})")
                skipped_count += 1
                continue
            
            # Get coordinates from nodes dictionary
            start_node = nodes.get(node_id_1)
            end_node = nodes.get(node_id_2)
            
            if not start_node or not end_node:
                print(f"Warning: Could not find nodes for connection {connection_id}")
                skipped_count += 1
                continue
            
            start_lat = start_node.get('latitude')
            start_lon = start_node.get('longitude')
            end_lat = end_node.get('latitude')
            end_lon = end_node.get('longitude')
            
            if any(coord is None for coord in [start_lat, start_lon, end_lat, end_lon]):
                print(f"Warning: Missing coordinates for connection {connection_id}")
                skipped_count += 1
                continue
            
            # Get connection type
            attributes = connection_data.get('attributes', {}).get('connection_type', {})
            connection_type = 'Unknown'
            if attributes.get('button_added'):
                connection_type = attributes.get('button_added')
            elif attributes.get('value'):
                connection_type = attributes.get('value')
            
            # Track connection types for analysis
            if connection_type not in connection_type_counts:
                connection_type_counts[connection_type] = 0
                connections_by_type[connection_type] = []
            connection_type_counts[connection_type] += 1
            connections_by_type[connection_type].append(connection_id)
            
            # Get attachment height if available
            attachment_height = None
            sections = connection_data.get('sections', {})
            for section_id, section_data in sections.items():
                if 'attachment_height' in section_data:
                    height = section_data['attachment_height']
                    if height is not None:
                        feet = int(height // 12)
                        inches = int(height % 12)
                        attachment_height = f"{feet}' {inches}\""
                        height_key = f"{feet}'{inches}\""
                        if height_key not in connection_height_counts:
                            connection_height_counts[height_key] = 0
                        connection_height_counts[height_key] += 1
                        connections_with_height += 1
                        break
            if attachment_height is None:
                connections_without_height += 1
            
            # Extract wire specification and mid-height from section photo data
            wire_spec = ''
            mid_height = None
            
            # Track if this is an aerial cable
            is_aerial_cable = connection_type == 'aerial cable'
            if is_aerial_cable:
                total_aerial_cables += 1
            
            # First, get all section photos
            for section_id, section_data in sections.items():
                section_photos = section_data.get('photos', {})
                for photo_id, photo_info in section_photos.items():
                    if photo_id in photo_data:
                        photo_data_item = photo_data[photo_id]
                        photofirst_data = photo_data_item.get('photofirst_data', {})
                        wire_data = photofirst_data.get('wire', {})
                        
                        # Look through each wire in the photo
                        for wire_info in wire_data.values():
                            trace_id = wire_info.get('_trace')
                            if trace_id and trace_id in trace_data:
                                trace_info = trace_data[trace_id]
                                
                                # Check if this is the Clearnetworx fiber optic wire
                                if (trace_info.get('company') == 'Clearnetworx' and
                                        trace_info.get('proposed', False) and
                                        trace_info.get('cable_type') == 'Fiber Optic Com'):
                                    wire_spec = wire_info.get('wire_spec', '')
                                    mid_height = wire_info.get('_measured_height')
                                    break
                        
                        if wire_spec and mid_height is not None:
                            break
            
            # Format mid-height
            mid_ht_str = ''
            if mid_height is not None:
                feet = int(mid_height // 12)
                inches = int(mid_height % 12)
                mid_ht_str = f"{feet}' {inches}\""
            
            # Track aerial cable stats
            if is_aerial_cable:
                if wire_spec:
                    aerial_with_wire_spec += 1
                if mid_ht_str:
                    aerial_with_mid_ht += 1
                
                # Store sample data for the first few aerial cables
                if len(sample_aerial_cables) < 5:
                    sample_aerial_cables.append({
                        'connection_id': connection_id,
                        'wire_spec': wire_spec,
                        'mid_ht': mid_ht_str
                    })
            
            # Create the line geometry and feature
            line = LineString([(start_lon, start_lat), (end_lon, end_lat)])
            
            # Get wire specification if available
            wire_spec = ''
            if 'wire_spec' in sections:
                wire_spec = sections['wire_spec']
            elif 'wire_specification' in sections:
                wire_spec = sections['wire_specification']

            feature = {
                'type': 'Feature',
                'geometry': mapping(line),
                'properties': {
                    'connection_id': connection_id,
                    'connection_type': connection_type,
                    'attachment_height': attachment_height,
                    'wire_spec': wire_spec,  # Add wire specification
                    'StartX': start_lon,
                    'StartY': start_lat,
                    'EndX': end_lon,
                    'EndY': end_lat,
                    'node_id_1': node_id_1,
                    'node_id_2': node_id_2
                }
            }
            
            valid_connections.append(feature)
            processed_count += 1
            
        except Exception as e:
            print(f"Error processing connection {connection_id}: {str(e)}")
            skipped_count += 1
            continue
    
    # Print aerial cable analysis
    print("\nAerial Cable Field Analysis:")
    print(f"Total aerial cables: {total_aerial_cables}")
    print(f"Aerial cables with wire_spec: {aerial_with_wire_spec} ({(aerial_with_wire_spec/total_aerial_cables*100 if total_aerial_cables > 0 else 0):.1f}%)")
    print(f"Aerial cables with mid_ht: {aerial_with_mid_ht} ({(aerial_with_mid_ht/total_aerial_cables*100 if total_aerial_cables > 0 else 0):.1f}%)")
    
    print("\nSample Aerial Cable Data:")
    for sample in sample_aerial_cables:
        print(f"Connection ID: {sample['connection_id']}")
        print(f"  wire_spec: {sample['wire_spec']}")
        print(f"  mid_ht: {sample['mid_ht']}")
        print("---")
    
    print(f"\nConnection Processing Summary:")
    print(f"Total connections: {total_connections}")
    print(f"Successfully processed: {processed_count}")
    print(f"Skipped: {skipped_count}")
    print("------------------------\n")
    
    return valid_connections

def savePointsToShapefile(points, filename):
    workspace_path = CONFIG['WORKSPACE_PATH']
    file_path = os.path.join(workspace_path, filename.replace('.shp', '.gpkg'))
    geometries = [Point(point["lng"], point["lat"]) for point in points]

    gdf = gpd.GeoDataFrame(points, geometry=geometries, crs="EPSG:4326")

    # Rename columns
    gdf.rename(columns={
        'company': 'utility',
        'tag': 'pole tag',
        'fldcompl': 'collected',
        'jobname': 'jobname',
        'job_status': 'job_status',
        'MR_statu': 'mr_status',
        'pole_spec': 'pole_spec',
        'POA_Height': 'att_ht',
        'lat': 'latitude',
        'lng': 'longitude'
    }, inplace=True)

    # Remove unwanted columns, ignore if they don't exist
    gdf.drop(columns=['pole_class', 'pole_height', 'id'], errors='ignore', inplace=True)

    # Save to file
    try:
        gdf.to_file(file_path, driver="GPKG")  # Switched to GeoPackage for better flexibility
        print(f"GeoPackage successfully saved to: {file_path}")
    except Exception as e:
        print(f"Error saving GeoPackage: {e}")


# Function to save line connections to a GeoPackage
def saveAnchorsToGeoPackage(anchor_points, filename):
    workspace_path = CONFIG['WORKSPACE_PATH']
    file_path = os.path.join(workspace_path, filename.replace('.shp', '.gpkg'))
    geometries = [Point(anchor["longitude"], anchor["latitude"]) for anchor in anchor_points]

    gdf = gpd.GeoDataFrame(anchor_points, geometry=geometries, crs="EPSG:4326")

    # Rename columns
    gdf.rename(columns={
        'longitude': 'longitude',
        'latitude': 'latitude',
        'anchor_spec': 'anchor_spec'
    }, inplace=True)

    # Save to file
    try:
        gdf.to_file(file_path, layer='anchors', driver="GPKG")
        print(f"Anchors GeoPackage successfully saved to: {file_path}")
    except Exception as e:
        print(f"Error saving anchors GeoPackage: {e}")
def saveLineShapefile(line_connections, filename):
    workspace_path = CONFIG['WORKSPACE_PATH']
    file_path = os.path.join(workspace_path, filename.replace('.shp', '.gpkg'))
    geometries = [
        LineString([(line["StartX"], line["StartY"]), (line["EndX"], line["EndY"])])
        for line in line_connections
    ]

    gdf = gpd.GeoDataFrame(line_connections, geometry=geometries, crs="EPSG:4326")
    gdf.drop(columns=['StartX', 'StartY', 'EndX', 'EndY', 'job_id'], errors='ignore', inplace=True)
    try:
        gdf.to_file(file_path, driver="GPKG")  # Switched to GeoPackage for better flexibility
        print(f"Line GeoPackage successfully saved to: {file_path}")
    except Exception as e:
        print(f"Error saving line GeoPackage: {e}")

# Function to save nodes to a GeoPackage
def saveMasterGeoPackage(all_nodes, all_connections, all_anchors, filename):
    workspace_path = CONFIG['WORKSPACE_PATH']
    file_path = os.path.join(workspace_path, filename)

    # Save poles as point layer (renamed from nodes)
    if all_nodes:
        try:
            # Create point geometries for poles
            geometries = [Point(node["lng"], node["lat"]) for node in all_nodes]
            gdf_poles = gpd.GeoDataFrame(all_nodes, geometry=geometries, crs="EPSG:4326")
            
            # Drop lat/lng columns as they're now in the geometry
            gdf_poles.drop(columns=['lat', 'lng'], errors='ignore', inplace=True)
            
            # Add project and conversation fields if not present
            if 'project' not in gdf_poles.columns:
                gdf_poles['project'] = ''
            if 'conversation' not in gdf_poles.columns:
                gdf_poles['conversation'] = ''
            
            # Save to GeoPackage
            gdf_poles.to_file(file_path, layer='poles', driver="GPKG")
            print(f"Poles layer successfully saved to: {file_path}")
            
        except Exception as e:
            print(f"Error saving poles layer to GeoPackage: {e}")

    # Save connections as line layer
    if all_connections:
        try:
            valid_connections = []
            line_geometries = []
            
            print(f"Processing {len(all_connections)} connections...")
            for connection in all_connections:
                try:
                    # Get coordinates from the connection properties
                    properties = connection.get('properties', {})
                    start_x = properties.get('StartX')
                    start_y = properties.get('StartY')
                    end_x = properties.get('EndX')
                    end_y = properties.get('EndY')

                    if any(coord is None for coord in [start_x, start_y, end_x, end_y]):
                        print(f"Missing coordinates for connection between nodes {properties.get('node_id_1')} and {properties.get('node_id_2')}")
                        continue

                    # Create LineString geometry
                    line_geom = LineString([(start_x, start_y), (end_x, end_y)])
                    
                    # Extract wire specification and mid-height from section photo data
                    section_photo_id = properties.get('section_photo_id')
                    if section_photo_id:
                        photo_data = connection.get('photo_data', {}).get(section_photo_id, {})
                        photofirst_data = photo_data.get('photofirst_data', {})
                        
                        # Extract mid-height
                        mid_height = photofirst_data.get('measured_height')
                        if mid_height is not None:
                            feet = int(mid_height // 12)
                            inches = int(mid_height % 12)
                            properties['mid_ht'] = f"{feet}' {inches}\""
                        
                        # Extract wire specification
                        wire_spec = photofirst_data.get('wire_spec', '')
                        if wire_spec:
                            properties['wire_spec'] = wire_spec
                    
                    # Use the existing properties
                    valid_connections.append(properties)
                    line_geometries.append(line_geom)
                        
                except Exception as e:
                    print(f"Error processing line: {str(e)}")
                    continue

            print(f"Found {len(valid_connections)} valid connections out of {len(all_connections)} total connections")
            
            if valid_connections and line_geometries:
                gdf_lines = gpd.GeoDataFrame(valid_connections, geometry=line_geometries, crs="EPSG:4326")
                
                # Rename 'attachment' to 'mid_ht' if it exists
                if 'attachment' in gdf_lines.columns:
                    gdf_lines.rename(columns={'attachment': 'mid_ht'}, inplace=True)
                
                if not gdf_lines.empty:
                    gdf_lines.to_file(file_path, layer='connections', driver="GPKG", mode='a')
                    print("Connections layer successfully saved to GeoPackage")
            else:
                print("No valid connections found")
                
        except Exception as e:
            print(f"Error saving connections layer to GeoPackage: {e}")

    # Save anchors as point layer
    if all_anchors:
        try:
            # Create point geometries for anchors
            geometries = [Point(anchor["longitude"], anchor["latitude"]) for anchor in all_anchors]
            gdf_anchors = gpd.GeoDataFrame(all_anchors, geometry=geometries, crs="EPSG:4326")
            
            # Drop lat/lng columns as they're now in the geometry
            gdf_anchors.drop(columns=['latitude', 'longitude'], errors='ignore', inplace=True)
            
            # Save to GeoPackage
            gdf_anchors.to_file(file_path, layer='anchors', driver="GPKG", mode='a')
            print("Anchors layer successfully saved to GeoPackage")
            
        except Exception as e:
            print(f"Error saving anchors layer to GeoPackage: {e}")

    print("Master GeoPackage saved successfully")

def update_sharepoint_spreadsheet(df, site_url=None, drive_path=None):
    """
    Update the spreadsheet in SharePoint with new data, supporting co-authoring
    """
    try:
        print("\nUpdating SharePoint spreadsheet...")
        
        # Debug info
        print(f"\nDataFrame Info:")
        print(f"Shape: {df.shape}")
        print(f"Columns: {list(df.columns)}")
        print(f"Sample of first row: {df.iloc[0].to_dict()}")
        
        # Sort DataFrame by Utility and Job Status
        df = df.sort_values(by=['Utility', 'Job Status'])
        
        # Format the data for SharePoint - ensure all values are strings
        formatted_data = []
        for _, row in df.iterrows():
            formatted_row = []
            for value in row:
                if pd.isna(value):
                    formatted_row.append("")
                else:
                    formatted_row.append(str(value))
            formatted_data.append(formatted_row)
        
        # Debug formatted data
        print(f"\nFormatted Data Info:")
        print(f"Number of rows: {len(formatted_data)}")
        if formatted_data:
            print(f"Sample formatted row: {formatted_data[0]}")
        
        # Use configured paths or fallback to parameters
        site_url = site_url or CONFIG['SHAREPOINT']['SITE_URL']
        drive_path = drive_path or CONFIG['SHAREPOINT']['DRIVE_PATH']
        file_name = CONFIG['SHAREPOINT']['FILE_NAME']
        file_path = f"{drive_path}/{file_name}"
        
        print(f"Using SharePoint path: {site_url}")
        print(f"Using file path: {file_path}")
        
        # Initialize the Graph client
        graph_client = initialize_graph_client()
        
        if not graph_client:
            print("Failed to initialize Graph client")
            return False
            
        # Get site ID
        print("Getting site ID...")
        site_response = graph_client.get(f"sites/{site_url}")
        if site_response.status_code != 200:
            print(f"Failed to get site. Status code: {site_response.status_code}")
            return False
            
        site_id = site_response.json()['id']
        print(f"Successfully got site ID: {site_id}")
        
        # Get drive ID
        print("Getting drive ID...")
        drives_response = graph_client.get(f"sites/{site_id}/drives")
        if drives_response.status_code != 200:
            print(f"Failed to get drives")
            return False
            
        # Find the Documents drive
        documents_drive = None
        for drive in drives_response.json()['value']:
            if drive['name'] == 'Documents':
                documents_drive = drive
                break
                
        if not documents_drive:
            print("Could not find Documents drive")
            return False
            
        drive_id = documents_drive['id']
        print(f"Successfully got drive ID: {drive_id}")
        
        # Get the file
        print(f"Checking for existing file at: {file_path}")
        file_response = graph_client.get(f"sites/{site_id}/drives/{drive_id}/root:/{file_path}")
        
        if file_response.status_code == 200:
            print("File exists, updating...")
            file_id = file_response.json()['id']
            
            # Create a workbook session
            session_response = graph_client.post(
                f"sites/{site_id}/drives/{drive_id}/items/{file_id}/workbook/createSession",
                json={"persistChanges": True}
            )
            
            if session_response.status_code != 201:
                print(f"Failed to create workbook session. Status code: {session_response.status_code}")
                return False
                
            session_id = session_response.json()['id']
            print("Successfully created workbook session")
            
            try:
                # Get the worksheet
                worksheet_response = graph_client.get(
                    f"sites/{site_id}/drives/{drive_id}/items/{file_id}/workbook/worksheets/Aerial%20Status%20Report",
                    headers={"workbook-session-id": session_id}
                )
                
                if worksheet_response.status_code != 200:
                    print(f"Failed to get worksheet. Status code: {worksheet_response.status_code}")
                    return False
                    
                # Update timestamp in row 2
                current_time = datetime.now().strftime('%-m/%-d/%Y %-I:%M %p MST')
                timestamp_response = graph_client.patch(
                    f"sites/{site_id}/drives/{drive_id}/items/{file_id}/workbook/worksheets/Aerial%20Status%20Report/range(address='A2:O2')",
                    headers={"workbook-session-id": session_id},
                    json={
                        "values": [[current_time] + [""] * 14]  # Empty strings for the rest of the row
                    }
                )
                
                if timestamp_response.status_code != 200:
                    print(f"Failed to update timestamp. Status code: {timestamp_response.status_code}")
                    
                # Clear existing data first (from row 4 onwards)
                clear_range = f"A4:O1000"  # Clear a large range to ensure all data is removed
                clear_response = graph_client.patch(
                    f"sites/{site_id}/drives/{drive_id}/items/{file_id}/workbook/worksheets/Aerial%20Status%20Report/range(address='{clear_range}')",
                    headers={"workbook-session-id": session_id},
                    json={
                        "values": [[""]*15]*997  # 997 rows of empty values (1000-3)
                    }
                )
                
                if clear_response.status_code != 200:
                    print(f"Failed to clear existing data. Status code: {clear_response.status_code}")
                    
                # Update data starting at A4 (preserving headers)
                if formatted_data:
                    num_rows = len(formatted_data)
                    data_range = f"A4:O{num_rows + 3}"  # +3 because we start at row 4 and Excel is 1-based
                    
                    # Get existing format
                    format_response = graph_client.get(
                        f"sites/{site_id}/drives/{drive_id}/items/{file_id}/workbook/worksheets/Aerial%20Status%20Report/range(address='{data_range}')/format",
                        headers={"workbook-session-id": session_id}
                    )
                    
                    if format_response.status_code == 200:
                        existing_format = format_response.json()
                        
                        # Debug print the exact data structure
                        print("\nDebug - SharePoint Update Data Structure:")
                        print(f"Data Range: {data_range}")
                        print("Update Payload:")
                        update_payload = {
                            "values": formatted_data
                        }
                        print(json.dumps(update_payload, indent=2))
                        
                        # Update values only, without format
                        update_response = graph_client.patch(
                            f"sites/{site_id}/drives/{drive_id}/items/{file_id}/workbook/worksheets/Aerial%20Status%20Report/range(address='{data_range}')",
                            headers={"workbook-session-id": session_id},
                            json=update_payload
                        )
                        
                        if update_response.status_code != 200:
                            print(f"Failed to update content. Status code: {update_response.status_code}")
                            print(f"Response: {update_response.text}")
                            return False
                    else:
                        print(f"Failed to get existing format. Status code: {format_response.status_code}")
                        return False
                        
                print("Successfully updated data while preserving formatting")
                return True
                
            finally:
                # Close the session
                try:
                    graph_client.post(
                        f"sites/{site_id}/drives/{drive_id}/items/{file_id}/workbook/closeSession",
                        headers={"workbook-session-id": session_id}
                    )
                except Exception as e:
                    print(f"Error closing session: {str(e)}")
                    
        else:
            print("File doesn't exist or cannot be accessed")
            return False
            
    except Exception as e:
        print(f"Error updating SharePoint spreadsheet: {str(e)}")
        return False

from excel_utils import create_summary_sheet

def create_report(jobs_summary):
    report_data = []

    for job in jobs_summary:
        job_name = job['job_name']
        job_status = job.get('job_status', 'Unknown').strip()
        mr_status_counts = job['mr_status_counts']
        pole_count = sum(mr_status_counts.values())
        
        # Get the fields from job summary
        field_complete_pct = job.get('field_complete_pct', 0)
        trace_complete_pct = job.get('trace_complete_pct', 0)
        utility = job.get('utility', 'Unknown')
        most_recent_editor = job.get('most_recent_editor', 'Unknown')
        last_edit_time = job.get('last_edit_time', 'Unknown')
        conversation = str(job.get('conversation', ''))  # Ensure conversation is a string
        project = job.get('project', '')
        assigned_osp = job.get('assigned_osp', 'Unknown')
        comments = job.get('comments', '')

        report_data.append({
            'Job Name': job_name,
            'Conversation': f"'{conversation}",  # Add leading apostrophe to force text format
            'Project': project,
            'Utility': utility,
            'Job Status': job_status,
            'Assigned OSP': assigned_osp,
            'Last Edit': last_edit_time,
            'Field %': f"{field_complete_pct:.1f}%",
            'Trace %': f"{trace_complete_pct:.1f}%",
            'No MR': mr_status_counts.get('No MR', 0),
            'Comm MR': mr_status_counts.get('Comm MR', 0),
            'Electric MR': mr_status_counts.get('Electric MR', 0),
            'PCO Required': mr_status_counts.get('PCO Required', 0),
            'Pole Count': pole_count,
            'Comments': comments
        })

    # Create a DataFrame from the report data
    df_report = pd.DataFrame(report_data)
    
    # Ensure Conversation column is treated as text
    df_report['Conversation'] = df_report['Conversation'].astype(str)

    # Sort the DataFrame first by Utility, then by Job Status
    df_report = df_report.sort_values(by=['Utility', 'Job Status'])

    # Ensure the directory exists
    workspace_dir = CONFIG['WORKSPACE_PATH']
    if not os.path.exists(workspace_dir):
        try:
            os.makedirs(workspace_dir)
            print(f"Workspace directory created: {workspace_dir}")
        except Exception as e:
            print(f"Failed to create workspace directory: {e}")
            return None

    # Generate a filename with a timestamp
    timestamp = datetime.now().strftime("%Y%m%d_%H%M")
    report_filename = f"Aerial_Status_Report_{timestamp}.xlsx"
    report_path = os.path.join(workspace_dir, report_filename)

    # Write the report to an Excel file with formatting
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Aerial Status Report"

        # Add merged header with title in the first row
        ws.merge_cells('A1:N1')
        title_cell = ws.cell(row=1, column=1)
        title_cell.value = "Aerial Status Report"
        title_cell.font = Font(size=18, bold=True)
        title_cell.alignment = Alignment(horizontal="center", vertical="center")

        # Add the date/time in the second row
        ws.merge_cells('A2:N2')
        date_cell = ws.cell(row=2, column=1)
        current_time = datetime.now()
        date_cell.value = current_time.strftime('%-m/%-d/%Y %-I:%M %p')
        date_cell.font = Font(size=12)
        date_cell.alignment = Alignment(horizontal="center", vertical="center")

        # Set row heights
        ws.row_dimensions[1].height = 30  # Title row
        ws.row_dimensions[2].height = 20  # Date row
        ws.row_dimensions[3].height = 25  # Column headers row

        # Add the column headers with styling in the third row
        column_widths = {
            "Job Name": 44,
            "Conversation": 15,
            "Project": 15,
            "Utility": 15,
            "Job Status": 23.71,
            "Assigned OSP": 30,
            "Last Edit": 20,
            "Field %": 10,
            "Trace %": 10,
            "No MR": 10,
            "Comm MR": 10,
            "Electric MR": 12,
            "PCO Required": 12,
            "Pole Count": 12,
            "Comments": 50  # Add width for Comments column
        }

        header_colors = {
            "Job Name": "CCFFCC",
            "Conversation": "CCFFCC",
            "Project": "CCFFCC",
            "Utility": "CCFFCC",
            "Job Status": "CCFFCC",
            "Assigned OSP": "CCFFCC",
            "Last Edit": "CCFFCC",
            "Field %": "CCFFCC",
            "Trace %": "CCFFCC",
            "No MR": "D9D9D9",
            "Comm MR": "FFFF00",
            "Electric MR": "FFC000",
            "PCO Required": "FF0000",
            "Pole Count": "CCFFCC",
            "Comments": "CCFFCC"  # Add color for Comments column
        }

        for col_num, column_title in enumerate(df_report.columns, 1):
            cell = ws.cell(row=3, column=col_num)
            cell.value = column_title
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Set column width
            column_letter = cell.column_letter
            ws.column_dimensions[column_letter].width = column_widths.get(column_title, 13.3)

            # Set header color
            if column_title in header_colors:
                cell.fill = PatternFill(start_color=header_colors[column_title],
                                     end_color=header_colors[column_title],
                                     fill_type="solid")

        # Add the data rows with center alignment
        for r_idx, row in enumerate(dataframe_to_rows(df_report, index=False, header=False), 4):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                cell.alignment = Alignment(horizontal="center", vertical="center")

        # Add borders around all cells
        thin_border = Border(left=Side(style='thin'),
                          right=Side(style='thin'),
                          top=Side(style='thin'),
                          bottom=Side(style='thin'))

        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=len(df_report.columns)):
            for cell in row:
                cell.border = thin_border

        # Save the workbook
        wb.save(report_path)
        print(f"Report successfully created: {report_path}")

    except Exception as e:
        print(f"Error creating report: {e}")
        return None

    try:
        # Update SharePoint spreadsheet
        print("\nUpdating SharePoint spreadsheet...")
        sharepoint_update_success = update_sharepoint_spreadsheet(
            df_report,
            CONFIG['SHAREPOINT']['SITE_URL'],
            CONFIG['SHAREPOINT']['DRIVE_PATH']
        )
        if sharepoint_update_success:
            print("SharePoint spreadsheet updated successfully")
        else:
            print("Failed to update SharePoint spreadsheet")
            
    except Exception as e:
        print(f"Error in SharePoint update: {str(e)}")

    return report_path

# Function to send email notification with attachment
def send_email_notification(recipients, report_path):
    """Send email notification with the report attached."""
    print("\nStarting email notification process...")
    print(f"Recipients: {recipients}")
    
    try:
        # Load environment variables
        load_dotenv()
        client_id = os.getenv('AZURE_CLIENT_ID')
        client_secret = os.getenv('AZURE_CLIENT_SECRET')
        tenant_id = os.getenv('AZURE_TENANT_ID')
        user_email = EMAIL_CONFIG['user_email']  # Use email from config instead of env var

        # Initialize MSAL client
        authority = f"https://login.microsoftonline.com/{tenant_id}"
        app = msal.ConfidentialClientApplication(
            client_id,
            authority=authority,
            client_credential=client_secret
        )

        # Get access token
        scopes = ['https://graph.microsoft.com/.default']
        result = app.acquire_token_silent(scopes, account=None)
        if not result:
            result = app.acquire_token_for_client(scopes)

        if 'access_token' in result:
            # SharePoint spreadsheet link
            sharepoint_link = "https://deeplydigital.sharepoint.com/:x:/s/OSPIntegrationTestingSite/EfNOio4T1RFEjmCOJ1XYb3YB2HZPT0v8Sqb--3A_mKJLSQ?e=kLEpux"
            
            # Prepare email message with SharePoint link
            email_msg = {
                'message': {
                    'subject': 'Aerial Status Report Generated',
                    'body': {
                        'contentType': 'Text',
                        'content': f'Please find attached the latest Aerial Status Report.\n\nYou can also view the report in SharePoint here:\n{sharepoint_link}\n\nThis is an automated email, if you would like to opt out please respond and I can remove you from the list.'
                    },
                    'toRecipients': [{'emailAddress': {'address': r}} for r in recipients],
                    'attachments': [{
                        '@odata.type': '#microsoft.graph.fileAttachment',
                        'name': os.path.basename(report_path),
                        'contentBytes': base64.b64encode(open(report_path, 'rb').read()).decode()
                    }]
                }
            }

            # Send email using Microsoft Graph API
            graph_endpoint = 'https://graph.microsoft.com/v1.0'
            headers = {
                'Authorization': f"Bearer {result['access_token']}",
                'Content-Type': 'application/json'
            }
            
            response = requests.post(
                f"{graph_endpoint}/users/{user_email}/sendMail",
                headers=headers,
                data=json.dumps(email_msg)
            )
            
            if response.status_code == 202:
                print("Email sent successfully")
            else:
                print(f"Failed to send email. Status code: {response.status_code}")
                print(f"Response: {response.text}")
        else:
            print(f"Error getting access token: {result.get('error_description')}")
            
    except Exception as e:
        print(f"Error in email notification: {str(e)}")
    
    print("Email notification process completed")

# Function to validate job data
def validateJobData(job_data):
    available_fields = []
    if 'nodes' in job_data:
        available_fields.append('nodes')
    if 'connections' in job_data:
        available_fields.append('connections')
    if 'metadata' in job_data:
        available_fields.append('metadata')
    
    print(f"Available fields in job data: {', '.join(available_fields)}")
    return True  # Always process the job with whatever data is available

def saveToShapefiles(nodes, connections, anchors, workspace_path):
    """Save nodes, connections, and anchors to shapefiles with WGS 1984 projection."""
    print("\nSaving data to shapefiles...")
    
    # Field name mappings (original -> truncated)
    node_fields = {
        'jobname': 'job_name',
        'job_status': 'job_stat',
        'MR_statu': 'mr_status',
        'company': 'utility',
        'fldcompl': 'completed',
        'tag': 'pole_tag',
        'POA_Height': 'poa_ht',
        'conversation': 'conv',  # New field
        'project': 'proj',  # New field
        'scid': 'scid',
        'last_editor': 'editor',
        'last_edit': 'edit_time',
        'node_id': 'node_id'  # New field for tracking
    }
    
    connection_fields = {
        'connection_id': 'conn_id',  # New field for tracking
        'connection_type': 'conn_type',
        'node_id_1': 'node1_id',
        'node_id_2': 'node2_id',
        'mid_ht': 'mid_ht',  # Renamed from attachment
        'wire_spec': 'wire_spec'  # New field
    }
    
    anchor_fields = {
        'anchor_spec': 'anch_spec',
        'job_id': 'job_id'
    }
    
    try:
        # Create a timestamp for the master zip file
        timestamp = datetime.now().strftime("%Y%m%d_%H%M")
        master_zip_name = f"KatapultMaster_{timestamp}.zip"
        master_zip_path = os.path.join(workspace_path, master_zip_name)
        
        shapefile_components = []

        # Save and analyze nodes
        if nodes:
            print("\nPole Type Analysis in Shapefile:")
            print("------------------------")
            node_geometries = [Point(node["lng"], node["lat"]) for node in nodes]
            gdf_nodes = gpd.GeoDataFrame(nodes, geometry=node_geometries, crs="EPSG:4326")
            
            # Analyze MR status distribution
            mr_status_counts = gdf_nodes['MR_statu'].value_counts()
            print("\nMR Status Distribution:")
            for status, count in mr_status_counts.items():
                print(f"{status}: {count}")
            
            # Add node_id if not present
            if 'node_id' not in gdf_nodes.columns:
                gdf_nodes['node_id'] = range(1, len(gdf_nodes) + 1)
            
            # Add project and conversation fields if not present
            if 'project' not in gdf_nodes.columns:
                gdf_nodes['project'] = ''
            if 'conversation' not in gdf_nodes.columns:
                gdf_nodes['conversation'] = ''
            
            # Rename columns and drop unnecessary ones
            gdf_nodes.rename(columns=node_fields, inplace=True)
            gdf_nodes.drop(columns=['lat', 'lng', 'pole_class', 'pole_height', 'id'], errors='ignore', inplace=True)
            
            nodes_shp = os.path.join(workspace_path, "poles.shp")
            gdf_nodes.to_file(nodes_shp, driver="ESRI Shapefile")
            print(f"\nPoles shapefile saved successfully with {len(gdf_nodes)} features")
            
            # Track shapefile components
            for ext in ['.shp', '.shx', '.dbf', '.prj', '.cpg']:
                file_path = nodes_shp.replace('.shp', ext)
                if os.path.exists(file_path):
                    shapefile_components.append(file_path)

        # Save and analyze connections
        if connections:
            print("\nConnection Type Analysis in Shapefile:")
            print("------------------------")
            valid_connections = []
            line_geometries = []
            connection_types = {}
            
            for connection in connections:
                try:
                    properties = connection.get('properties', {})
                    conn_type = properties.get('connection_type', 'Unknown')
                    if conn_type not in connection_types:
                        connection_types[conn_type] = 0
                    connection_types[conn_type] += 1
                    
                    start_x = properties.get('StartX')
                    start_y = properties.get('StartY')
                    end_x = properties.get('EndX')
                    end_y = properties.get('EndY')
                    
                    if any(coord is None for coord in [start_x, start_y, end_x, end_y]):
                        continue
                        
                    line_geom = LineString([(start_x, start_y), (end_x, end_y)])
                    
                    # Add connection_id if not present
                    if 'connection_id' not in properties:
                        properties['connection_id'] = len(valid_connections) + 1
                    
                    # Extract wire specification and mid-height from section photo data
                    section_photo_id = properties.get('section_photo_id')
                    if section_photo_id:
                        photo_data = connection.get('photo_data', {}).get(section_photo_id, {})
                        photofirst_data = photo_data.get('photofirst_data', {})
                        
                        # Extract mid-height
                        mid_height = photofirst_data.get('measured_height')
                        if mid_height is not None:
                            feet = int(mid_height // 12)
                            inches = int(mid_height % 12)
                            properties['mid_ht'] = f"{feet}' {inches}\""
                        
                        # Extract wire specification
                        wire_spec = photofirst_data.get('wire_spec', '')
                        if wire_spec:
                            properties['wire_spec'] = wire_spec
                    
                    valid_connections.append(properties)
                    line_geometries.append(line_geom)
                    
                except Exception as e:
                    continue

            # Print connection type counts
            for conn_type, count in sorted(connection_types.items()):
                print(f"{conn_type}: {count}")
            
            if valid_connections and line_geometries:
                gdf_connections = gpd.GeoDataFrame(valid_connections, geometry=line_geometries, crs="EPSG:4326")
                
                # Rename 'attachment' to 'mid_ht' if it exists
                if 'attachment' in gdf_connections.columns:
                    gdf_connections.rename(columns={'attachment': 'mid_ht'}, inplace=True)
                
                gdf_connections.rename(columns=connection_fields, inplace=True)
                
                connections_shp = os.path.join(workspace_path, "connections.shp")
                gdf_connections.to_file(connections_shp, driver="ESRI Shapefile")
                print(f"\nConnections shapefile saved successfully with {len(gdf_connections)} features")
                
                # Track shapefile components
                for ext in ['.shp', '.shx', '.dbf', '.prj', '.cpg']:
                    file_path = connections_shp.replace('.shp', ext)
                    if os.path.exists(file_path):
                        shapefile_components.append(file_path)
    
        # Save and analyze anchors
        if anchors:
            print("\nAnchor Type Analysis in Shapefile:")
            print("------------------------")
            anchor_geometries = [Point(anchor["longitude"], anchor["latitude"]) for anchor in anchors]
            gdf_anchors = gpd.GeoDataFrame(anchors, geometry=anchor_geometries, crs="EPSG:4326")
            
            # Analyze anchor spec distribution
            anchor_spec_counts = gdf_anchors['anchor_spec'].value_counts()
            print("\nAnchor Spec Distribution:")
            for spec, count in anchor_spec_counts.items():
                print(f"{spec}: {count}")
            
            gdf_anchors.rename(columns=anchor_fields, inplace=True)
            gdf_anchors.drop(columns=['latitude', 'longitude'], errors='ignore', inplace=True)
            
            anchors_shp = os.path.join(workspace_path, "anchors.shp")
            gdf_anchors.to_file(anchors_shp, driver="ESRI Shapefile")
            print(f"\nAnchors shapefile saved successfully with {len(gdf_anchors)} features")
            
            # Track shapefile components
            for ext in ['.shp', '.shx', '.dbf', '.prj', '.cpg']:
                file_path = anchors_shp.replace('.shp', ext)
                if os.path.exists(file_path):
                    shapefile_components.append(file_path)

        # Create master zip file containing all shapefile components
        with zipfile.ZipFile(master_zip_path, 'w') as master_zip:
            for file_path in shapefile_components:
                if os.path.exists(file_path):
                    master_zip.write(file_path, os.path.basename(file_path))
                    os.remove(file_path)  # Remove the original file after adding to zip
        
        print(f"\nAll shapefiles have been consolidated into: {master_zip_path}")
        
        # Upload master zip to SharePoint
        print("\nUploading master zip to SharePoint...")
        graph_client = initialize_graph_client()
        if not graph_client:
            print("Failed to initialize Graph client")
            return
            
        # Get site ID
        site_url = CONFIG['SHAREPOINT']['SITE_URL']
        site_response = graph_client.get(f"sites/{site_url}")
        if site_response.status_code != 200:
            print(f"Failed to get site. Status code: {site_response.status_code}")
            return
            
        site_id = site_response.json()['id']
        
        # Get drive ID
        drives_response = graph_client.get(f"sites/{site_id}/drives")
        if drives_response.status_code != 200:
            print(f"Failed to get drives")
            return
            
        # Find the Documents drive
        documents_drive = None
        for drive in drives_response.json()['value']:
            if drive['name'] == 'Documents':
                documents_drive = drive
                break

        if not documents_drive:
            print("Could not find Documents drive")
            return
            
        drive_id = documents_drive['id']
        
        # Check for existing zip files and delete them
        print("Checking for existing shapefile zip files...")
        files_response = graph_client.get(f"sites/{site_id}/drives/{drive_id}/root/children")
        if files_response.status_code == 200:
            for item in files_response.json().get('value', []):
                if item['name'].startswith('KatapultMaster_') and item['name'].endswith('.zip'):
                    print(f"Deleting existing file: {item['name']}")
                    delete_response = graph_client.delete(f"sites/{site_id}/drives/{drive_id}/items/{item['id']}")
                    if delete_response.status_code != 204:
                        print(f"Failed to delete file {item['name']}")
        
        # Upload the new zip file
        print(f"Uploading new file: {master_zip_name}")
        with open(master_zip_path, 'rb') as file_content:
            # Create upload session for Documents folder
            upload_session = graph_client.post(
                f"sites/{site_id}/drives/{drive_id}/root:/{master_zip_name}:/createUploadSession",
                json={
                    "@microsoft.graph.conflictBehavior": "replace"
                }
            )
            
            if upload_session.status_code != 200:
                print("Failed to create upload session")
                return
                
            upload_url = upload_session.json()['uploadUrl']
            
            # Read file content
            file_content.seek(0, 2)  # Seek to end
            file_size = file_content.tell()
            file_content.seek(0)  # Seek back to start
            
            # Upload the file in chunks
            chunk_size = 320 * 1024  # 320 KB chunks
            for chunk_start in range(0, file_size, chunk_size):
                chunk_end = min(chunk_start + chunk_size - 1, file_size - 1)
                content_length = chunk_end - chunk_start + 1
                
                file_content.seek(chunk_start)
                chunk_data = file_content.read(content_length)
                
                headers = {
                    'Content-Length': str(content_length),
                    'Content-Range': f'bytes {chunk_start}-{chunk_end}/{file_size}'
                }
                
                chunk_response = requests.put(upload_url, data=chunk_data, headers=headers)
                if chunk_response.status_code not in [200, 201, 202]:
                    print(f"Failed to upload chunk. Status code: {chunk_response.status_code}")
                    return
        
        print("Master zip file successfully uploaded to SharePoint")
            
    except Exception as e:
        print(f"Error saving shapefiles: {str(e)}")
        
    print("\nVerification Summary:")
    print("------------------------")
    print(f"Input Counts:")
    print(f"Poles: {len(nodes)}")
    print(f"Connections: {len(connections)}")
    print(f"Anchors: {len(anchors)}")
    print("------------------------")

def test_sharepoint_access():
    """Test SharePoint access using existing credentials."""
    try:
        # Load environment variables
        load_dotenv()
        client_id = os.getenv('AZURE_CLIENT_ID')
        client_secret = os.getenv('AZURE_CLIENT_SECRET')
        tenant_id = os.getenv('AZURE_TENANT_ID')
        user_email = os.getenv('EMAIL_USER')

        # Initialize MSAL client
        authority = f"https://login.microsoftonline.com/{tenant_id}"
        app = msal.ConfidentialClientApplication(
            client_id,
            authority=authority,
            client_credential=client_secret
        )

        # Get access token with SharePoint scope
        scopes = ['https://graph.microsoft.com/.default']
        result = app.acquire_token_silent(scopes, account=None)
        if not result:
            result = app.acquire_token_for_client(scopes)

        if 'access_token' in result:
            # Test SharePoint access using Microsoft Graph API
            headers = {
                'Authorization': f"Bearer {result['access_token']}",
                'Content-Type': 'application/json'
            }
            
            # Make a test request to SharePoint
            graph_endpoint = 'https://graph.microsoft.com/v1.0'
            response = requests.get(
                f"{graph_endpoint}/sites",
                headers=headers
            )
            
            if response.status_code == 200:
                print("Successfully connected to SharePoint")
                return True
            else:
                print(f"Failed to access SharePoint. Status code: {response.status_code}")
                print(f"Response: {response.text}")
                return False
        else:
            print(f"Error getting access token: {result.get('error_description')}")
            return False
            
    except Exception as e:
        print(f"Error testing SharePoint access: {str(e)}")
        return False

def initialize_graph_client():
    """Initialize and return a Microsoft Graph API client"""
    try:
        # Load environment variables
        load_dotenv()
        client_id = os.getenv('AZURE_CLIENT_ID')
        client_secret = os.getenv('AZURE_CLIENT_SECRET')
        tenant_id = os.getenv('AZURE_TENANT_ID')
        
        # Initialize MSAL client
        authority = f"https://login.microsoftonline.com/{tenant_id}"
        app = msal.ConfidentialClientApplication(
            client_id,
            authority=authority,
            client_credential=client_secret
        )
        
        # Get access token
        scopes = ['https://graph.microsoft.com/.default']
        result = app.acquire_token_silent(scopes, account=None)
        if not result:
            result = app.acquire_token_for_client(scopes)
            
        if 'access_token' not in result:
            print(f"Error getting access token: {result.get('error_description')}")
            return None
            
        # Create a requests Session with the token
        session = requests.Session()
        session.headers.update({
            'Authorization': f"Bearer {result['access_token']}",
            'Accept': 'application/json'
        })
        session.base_url = 'https://graph.microsoft.com/v1.0'
        
        # Add method to handle full URLs
        def request_with_base_url(method, url, **kwargs):
            if not url.startswith('http'):
                url = f"{session.base_url}/{url}"
            return session.request(method, url, **kwargs)
            
        session.get = lambda url, **kwargs: request_with_base_url('GET', url, **kwargs)
        session.post = lambda url, **kwargs: request_with_base_url('POST', url, **kwargs)
        session.put = lambda url, **kwargs: request_with_base_url('PUT', url, **kwargs)
        session.patch = lambda url, **kwargs: request_with_base_url('PATCH', url, **kwargs)
        session.delete = lambda url, **kwargs: request_with_base_url('DELETE', url, **kwargs)
        
        return session
        
    except Exception as e:
        print(f"Error initializing Graph client: {str(e)}")
        return None

# Function to get user list from KatapultPro API
def getUserList():
    URL_PATH = '/api/v2/users'
    headers = {}
    user_map = {}

    for attempt in range(5):
        conn = None
        try:
            conn = http.client.HTTPSConnection("katapultpro.com", timeout=10)
            conn.request("GET", f"{URL_PATH}?api_key={CONFIG['API_KEY']}", headers=headers)
            res = conn.getresponse()
            data = res.read().decode("utf-8")
            users_dict = json.loads(data)

            if not isinstance(users_dict, dict):
                raise TypeError(f"Expected a dictionary but received {type(users_dict)}: {users_dict}")

            # Create a mapping of user IDs to full names
            for user_id, user_data in users_dict.items():
                name = user_data.get('name', {})
                full_name = f"{name.get('first', '')} {name.get('last', '')}".strip()
                if not full_name:
                    full_name = user_data.get('email', 'Unknown User')
                user_map[user_id] = full_name

            logging.info(f"Retrieved {len(user_map)} users")
            break

        except (socket.error, OSError) as e:
            print(f"Socket error while getting user list: {e}. Retrying...")
            time.sleep(5)
        except Exception as e:
            print(f"Failed to retrieve user list: {e}")
            break
        finally:
            if conn:
                conn.close()

    return user_map

def update_arcgis_features(nodes, connections, anchors):
    """Update ArcGIS feature services with the latest data"""
    try:
        logging.info("\nUpdating ArcGIS feature services...")
        updater = ArcGISUpdater()
        
        # Process poles
        if nodes:
            pole_features = []
            for node in nodes:
                feature = {
                    'geometry': {
                        'x': node['lng'],
                        'y': node['lat'],
                        'spatialReference': {'wkid': 4326}
                    },
                    'attributes': {
                        'node_id': node.get('node_id', ''),
                        'job_name': node.get('jobname', ''),
                        'job_status': node.get('job_status', ''),
                        'mr_status': node.get('MR_statu', ''),
                        'utility': node.get('company', ''),
                        'completed': node.get('fldcompl', ''),
                        'pole_tag': node.get('tag', ''),
                        'poa_ht': node.get('POA_Height', ''),
                        'last_editor': node.get('last_editor', ''),
                        'last_edit': node.get('last_edit', '')
                    }
                }
                pole_features.append(feature)
            if not updater.update_features('poles', pole_features):
                logging.error("Failed to update poles feature service")
                return False

        # Process connections
        if connections:
            connection_features = []
            for conn in connections:
                properties = conn.get('properties', {})
                feature = {
                    'geometry': {
                        'paths': [[
                            [conn['geometry']['coordinates'][0][0], conn['geometry']['coordinates'][0][1]],
                            [conn['geometry']['coordinates'][1][0], conn['geometry']['coordinates'][1][1]]
                        ]],
                        'spatialReference': {'wkid': 4326}
                    },
                    'attributes': {
                        'conn_id': properties.get('connection_id', ''),
                        'conn_type': properties.get('connection_type', ''),
                        'att_height': properties.get('attachment_height', ''),
                        'node_id_1': properties.get('node_id_1', ''),
                        'node_id_2': properties.get('node_id_2', ''),
                        'wire_spec': properties.get('wire_spec', '')  # Add wire_spec field
                    }
                }
                connection_features.append(feature)
            if not updater.update_features('connections', connection_features):
                logging.error("Failed to update connections feature service")
                return False

        # Process anchors
        if anchors:
            anchor_features = []
            for anchor in anchors:
                feature = {
                    'geometry': {
                        'x': anchor['longitude'],
                        'y': anchor['latitude'],
                        'spatialReference': {'wkid': 4326}
                    },
                    'attributes': {
                        'anch_spec': anchor.get('anchor_spec', ''),
                        'job_id': anchor.get('job_id', ''),
                        'anchor_type': anchor.get('anchor_type', '')
                    }
                }
                anchor_features.append(feature)
            if not updater.update_features('anchors', anchor_features):
                logging.error("Failed to update anchors feature service")
                return False

        logging.info("Successfully updated all ArcGIS feature services")
        return True

    except Exception as e:
        logging.error(f"Error updating ArcGIS feature services: {str(e)}")
        logging.error(f"Stack trace: {traceback.format_exc()}")
        return False

# Main function to run the job for testing
def main(email_list):
    """Main function to process jobs and generate reports."""
    print("Starting main function...")
    all_jobs = []
    
    # Get user list first
    print("Getting user list...")
    user_map = getUserList()
    print(f"Retrieved {len(user_map)} users")
    
    if TEST_ONLY_SPECIFIC_JOB:
        print(f"Testing specific job with ID: {TEST_JOB_IDS}")
        # Get job list to find actual job names
        try:
            conn = http.client.HTTPSConnection("katapultpro.com", timeout=30)
            conn.request("GET", f"/api/v2/jobs?api_key={CONFIG['API_KEY']}")
            res = conn.getresponse()
            data = res.read().decode("utf-8")
            all_jobs_dict = json.loads(data)
            
            # Create job list with actual names
            all_jobs = []
            for job_id in TEST_JOB_IDS:
                job_data = all_jobs_dict.get(job_id, {})
                # Get job name from metadata or fallback to name field
                metadata = job_data.get('metadata', {})
                job_name = metadata.get('name')
                if not job_name:
                    job_name = job_data.get('name')
                if not job_name:  # If still no name found, use job ID
                    job_name = f"Job {job_id}"
                all_jobs.append({'id': job_id, 'name': job_name})
                print(f"Job ID: {job_id}, Name: {job_name}")
        except Exception as e:
            print(f"Error getting job names: {e}")
            # If we fail to get names, create jobs list with IDs as names
            all_jobs = [{'id': job_id, 'name': f"Job {job_id}"} for job_id in TEST_JOB_IDS]
            if 'data' in locals():
                print(f"Response data: {data[:200]}...")  # Print first 200 chars of response
        finally:
            if conn:
                conn.close()
    else:
        print("Getting list of all jobs...")
        all_jobs = getJobList()
    
    # Initialize lists to store all data
    all_nodes = []
    all_connections = []
    all_anchors = []
    jobs_summary = []
    
    if not all_jobs:
        print("No jobs found.")
        return False
    
    total_jobs = len(all_jobs)
    print(f"Found {total_jobs} jobs to process")
    
    # Test SharePoint access first
    print("\nTesting SharePoint access...")
    sharepoint_access = test_sharepoint_access()
    if not sharepoint_access:
        print("Warning: Could not access SharePoint. Will generate local report only.")
    
    for index, job in enumerate(all_jobs, 1):
        print(f"\n{'='*50}")
        print(f"Processing job {index}/{total_jobs}: {job['name']}")
        print(f"{'='*50}")
        
        job_id = job['id']
        job_name = job['name']
        
        print(f"Fetching data for job: {job_name} (ID: {job_id})")
        job_data = getJobData(job_id)
        
        if job_data and validateJobData(job_data):
            # Extract nodes and connections from job data
            nodes = job_data.get('nodes', {})
            connections = job_data.get('connections', {})
            metadata = job_data.get('metadata', {})
            
            print("\nDEBUG - Metadata Contents:")
            print(f"Raw metadata: {json.dumps(metadata, indent=2)}")
            
            # Update job name from metadata if available
            if metadata:
                metadata_name = metadata.get('name')
                if metadata_name:
                    job_name = metadata_name
                    print(f"DEBUG - Using job name from metadata: {job_name}")
                else:
                    # Try to get name from job data directly
                    job_data_name = job_data.get('name')
                    if job_data_name:
                        job_name = job_data_name
                        print(f"DEBUG - Using name from job data: {job_name}")
                    else:
                        print(f"DEBUG - Using existing job name: {job_name}")
            
            print("\nExtracting nodes...")
            nodes_data = extractNodes(job_data, job_name, job_id, user_map)
            print(f"Found {len(nodes_data)} nodes")
            
            print("\nExtracting connections...")
            connections_data = extractConnections(connections, job_data.get('nodes', {}), job_data)
            
            print("\nExtracting anchors...")
            anchors = extractAnchors(job_data, job_name, job_id)
            print(f"Found {len(anchors)} anchors")
            
            if nodes_data:
                print("\nDEBUG - Processing nodes for job summary...")
                all_nodes.extend(nodes_data)
                
                # Calculate field completion percentage for poles only
                poles = {node_id: node_data for node_id, node_data in nodes.items()
                        if any(node_data.get('attributes', {}).get(type_field, {}).get(source) == 'pole'
                            for type_field in ['node_type', 'pole_type']
                            for source in ['button_added', '-Imported', 'value', 'auto_calced'])}
                
                total_poles = len(poles)
                field_completed = sum(1 for node in poles.values() 
                                   if node.get('attributes', {}).get('field_completed', {}).get('value') == 1)
                field_complete_pct = (field_completed / total_poles * 100) if total_poles > 0 else 0
                
                print(f"\nDEBUG - Field Completion Stats:")
                print(f"Total poles found: {total_poles}")
                print(f"Field completed poles: {field_completed}")
                print(f"Field completion percentage: {field_complete_pct:.2f}%")
                
                # Calculate trace completion percentage by checking only aerial cable connections
                aerial_cable_connections = {
                    conn_id: conn_data for conn_id, conn_data in connections.items()
                    if conn_data.get('attributes', {}).get('connection_type', {}).get('button_added') == 'aerial cable' or
                       conn_data.get('attributes', {}).get('connection_type', {}).get('value') == 'aerial cable'
                }
                total_traces = len(aerial_cable_connections)
                completed_traces = sum(1 for connection in aerial_cable_connections.values() 
                                      if connection.get('attributes', {}).get('tracing_complete', {}).get('auto', False))
                trace_complete_pct = (completed_traces / total_traces * 100) if total_traces > 0 else 0
                
                print(f"\nDEBUG - Trace Completion Stats:")
                print(f"Total aerial cable connections: {total_traces}")
                print(f"Completed traces: {completed_traces}")
                print(f"Trace completion percentage: {trace_complete_pct:.2f}%")
                
                # Find most recent editor and edit time
                most_recent_editor = 'Unknown'
                last_edit_time = 'Unknown'
                latest_timestamp = 0
                
                print("\nDEBUG - Processing edit timestamps...")
                for node in nodes_data:
                    if node.get('last_editor') and node.get('last_edit'):
                        print(f"Found edit: Editor={node['last_editor']}, Time={node['last_edit']}")
                        try:
                            edit_time = node['last_edit'].replace(' MST', '')
                            edit_dt = datetime.strptime(edit_time, '%Y-%m-%d %I:%M %p')
                            timestamp = edit_dt.timestamp()
                            
                            if timestamp > latest_timestamp:
                                latest_timestamp = timestamp
                                most_recent_editor = node['last_editor']
                                last_edit_time = node['last_edit']
                                print(f"New most recent edit found: {last_edit_time} by {most_recent_editor}")
                        except Exception as e:
                            print(f"Error parsing edit time: {e}")
                            continue
                
                # Get utility from first pole with a company value
                utility = 'Unknown'
                print("\nDEBUG - Searching for utility company...")
                for node_id, node_data in job_data.get('nodes', {}).items():
                    pole_tag = node_data.get('attributes', {}).get('pole_tag', {})
                    print(f"\nChecking node {node_id}:")
                    print(f"Pole tag data: {json.dumps(pole_tag, indent=2)}")
                    
                    # First try to get company from -Imported or button_added
                    company = pole_tag.get('-Imported', {}).get('company') or pole_tag.get('button_added', {}).get('company')
                    
                    # If not found, look for company in any direct child of pole_tag
                    if not company:
                        for key, value in pole_tag.items():
                            if isinstance(value, dict) and 'company' in value:
                                company = value['company']
                                print(f"Found company in {key}: {company}")
                                break
                    
                    if company:
                        utility = company
                        print(f"Using utility: {utility}")
                        break
                
                # Get job status and assigned OSP from metadata
                job_status = metadata.get('job_status', 'Unknown')
                assigned_osp = metadata.get('assigned_OSP', 'Unknown')
                conversation = metadata.get('conversation', '')
                project = metadata.get('project', '')
                comments = metadata.get('comments', '')  # Extract comments from metadata
                
                print("\nDEBUG - Metadata fields:")
                print(f"Job Status: {job_status}")
                print(f"Assigned OSP: {assigned_osp}")
                print(f"Conversation: {conversation}")
                print(f"Project: {project}")
                print(f"Comments: {comments}")
                
                # Summarize MR Status counts for the job
                mr_status_counts = {}
                for node in nodes_data:
                    mr_status = node['MR_statu']
                    if mr_status not in mr_status_counts:
                        mr_status_counts[mr_status] = 0
                    mr_status_counts[mr_status] += 1
                
                print("\nDEBUG - MR Status Counts:")
                print(json.dumps(mr_status_counts, indent=2))
                
                jobs_summary.append({
                    'job_name': job_name,
                    'job_status': job_status,
                    'mr_status_counts': mr_status_counts,
                    'field_complete_pct': field_complete_pct,
                    'trace_complete_pct': trace_complete_pct,
                    'utility': utility,
                    'most_recent_editor': most_recent_editor,
                    'last_edit_time': last_edit_time,
                    'assigned_osp': assigned_osp,
                    'conversation': conversation,
                    'project': project,
                    'comments': comments  # Add comments to job summary
                })
                
                print("\nDEBUG - Final job summary entry:")
                print(json.dumps(jobs_summary[-1], indent=2))
                
                if connections_data:
                    all_connections.extend(connections_data)
                if anchors:
                    all_anchors.extend(anchors)
            else:
                print("No valid nodes found for job")
        else:
            print(f"Invalid or missing data for job {job_name}")
            continue
            
        print(f"Completed processing job {index}/{total_jobs}")
        
    print("\nAll jobs processed")
    print(f"Total nodes: {len(all_nodes)}")
    print(f"Total connections: {len(all_connections)}")
    print(f"Total anchors: {len(all_anchors)}")
    
    if all_nodes or all_connections or all_anchors:
        workspace_path = CONFIG['WORKSPACE_PATH']
        
        # Update ArcGIS feature services
        print("\nUpdating ArcGIS feature services...")
        arcgis_success = update_arcgis_features(all_nodes, all_connections, all_anchors)
        if arcgis_success:
            print("ArcGIS feature services updated successfully")
        else:
            print("Failed to update ArcGIS feature services")

        # Save to shapefiles and update SharePoint
        print("\nSaving data to shapefiles...")
        saveToShapefiles(all_nodes, all_connections, all_anchors, workspace_path)
    else:
        print("No data extracted for any job. Nothing to save.")
    
    print("\nGenerating report...")
    report_path = None
    if jobs_summary:
        report_path = create_report(jobs_summary)
        print(f"Report generated at: {report_path}")
        
        print("\nSending email notification...")
        send_email_notification(email_list, report_path)
        print("Email notification process completed")
    else:
        print("No job summary data available. Report not generated.")
    
    print("\nMain function completed")
    return True

def run_job():
    try:
        # Put your main script logic here
        logging.info("==============================================")
        logging.info("STARTING KATAPULT AUTOMATION JOB")
        logging.info("==============================================")
        email_list = [
            "brandan.lewis@deeplydigital.com"
        ]
        result = main(email_list)
        if result:
            logging.info("Job completed successfully")
            return {"status": "success", "message": "Job completed successfully"}
        else:
            logging.error("Job failed")
            return {"status": "error", "message": "Job failed"}
    except Exception as e:
        logging.error(f"Error in job execution: {str(e)}")
        return {"status": "error", "message": str(e)}

@app.route('/', methods=['POST'])
def handle_request():
    """Handle incoming HTTP POST requests"""
    try:
        result = run_job()
        return jsonify(result)
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500

@app.route('/', methods=['GET'])
def health_check():
    """Handle health check requests"""
    return jsonify({"status": "healthy"}), 200

if __name__ == '__main__':
    # Get port from environment variable or default to 8080
    port = int(os.environ.get('PORT', 8080))
    # Run the Flask app
    app.run(host='0.0.0.0', port=port)


